import openpyxl

def transpose_excel(input_file, output_file):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active

    # Get the number of rows and columns
    num_rows = sheet.max_row
    num_cols = sheet.max_column

    # Create a new workbook for the transposed data
    transposed_wb = openpyxl.Workbook()
    transposed_sheet = transposed_wb.active

    # Transpose the data
    for row in range(1, num_rows + 1):
        for col in range(1, num_cols + 1):
            transposed_sheet.cell(row=col, column=row, value=sheet.cell(row=row, column=col).value)

    # Save the transposed data to the output file
    transposed_wb.save(output_file)

# Example usage
input_file_path = 'test5000rows.xlsx'
output_file_path = 'output2.xlsx'
transpose_excel(input_file_path, output_file_path)
